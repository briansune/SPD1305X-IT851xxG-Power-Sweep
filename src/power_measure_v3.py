from __future__ import print_function
import builtins
import spd1000_series as _spd
import it85xxg_series as _it8
import math
from datetime import datetime
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference, LineChart, Series
import os
from time import sleep
import concurrent.futures
import logging

sys_ver = 0


def psu_on(spd_hdr):
    while True:
        spd_hdr.spdSetup(b'OUTP CH1,ON')
        res = spd_hdr.spdQuery(b'SYST:STAT?')
        if int(res, 16) & 0x10 == 0x10:
            break


def psu_off(spd_hdr):
    while True:
        spd_hdr.spdSetup(b'OUTP CH1,OFF')
        res = spd_hdr.spdQuery(b'SYST:STAT?')
        if int(res, 16) & 0x10 == 0x00:
            break


def pld_acq(pld_hdr):
    pld_hdr.it85_get_coarse_measure()
    prg_load = [
        pld_hdr.it85_mes_dict['Volt Avg'],
        pld_hdr.it85_mes_dict['Curr Avg'],
        pld_hdr.it85_mes_dict['Powr Avg']]
    return prg_load


def psu_acq(spd_hdr):
    psu_load = [
        spd_hdr.spdQuery(b'MEAS:VOLT?'),
        spd_hdr.spdQuery(b'MEAS:CURR?'),
        spd_hdr.spdQuery(b'MEAS:POWE?')]
    return psu_load


def main():
    # Logging entry
    log_format = logging.Formatter('%(asctime)s: %(levelname)s: %(message)s')
    logger = logging.getLogger('psu_pld_logger')
    logger.setLevel(logging.INFO)
    # print = logger.info
    _print = print  # keep a local copy of the original print
    builtins.print = logger.info

    pld_hdr = _it8.IT85xxPlus()
    pld_hdr.it85_ip = "192.168.2.212"

    spd_hdr = _spd.SPD1000()
    spd_hdr.spd_ip = "192.168.2.214"

    global sys_ver

    sys_ver = sys.version_info[0]
    print(sys_ver)

    spd_hdr.spdConnect()
    pld_hdr.it85_connect()

    dt = datetime.now()
    ts = math.floor(datetime.timestamp(dt))

    # Turn off programmable load
    pld_hdr.it85_channel_on_off(False)

    # ==============================================================================
    # Change the following parameters!
    # ==============================================================================
    spd_hdr.spdSetup(b'*UNLOCK')
    # spd_hdr.spdSetup(b'MODE:SET 2W')
    spd_hdr.spdSetup(b'MODE:SET 4W')

    sweep_volt_low = 28.0
    sweep_volt_high = 29.0
    sweep_step = 1.0
    # Repeat Num set
    repeat = 1
    # Other offset power
    offset = 0.000

    targ_max_curr = 0.3
    targ_max_volt = 5.05
    sweep_curr_step = 100

    pl_max_curr = targ_max_curr * 1.3
    pl_max_volt = targ_max_volt * 1.3
    pl_max_curr = pl_max_curr if pl_max_curr < 30.0 else 30.0
    pl_max_pwr = pl_max_curr * pl_max_volt
    pl_4wire_sen = True
    pl_cycle_step = False

    sweep_curr_max = int(targ_max_curr * 1000.0)
    psu_curr = (sweep_curr_max / 1000 * pl_max_volt) / sweep_volt_low * 1.2
    psu_curr = 5.0 if psu_curr > 5.0 else psu_curr

    # filename = 'TPS65261_5Vto3v3_2A_HF_1U5H'
    # filename = 'MP2145_5to1V_470nH'
    filename = 'MPQ4470G_30to5V_4u7H'

    user_msg = 'F=100k, RAMP2=150k,1n5F,330nF, FB=51k, 10k2'

    fh = logging.FileHandler('./log/{}_{}.log'.format(filename, ts))
    fh.setFormatter(log_format)
    fh.setLevel(logging.INFO)
    logger.addHandler(fh)

    ch = logging.StreamHandler()
    ch.setFormatter(log_format)
    logger.addHandler(ch)

    filename = filename + '.xlsx'
    print(filename)

    print("Power Supply Voltage Range: {} to {}".format(sweep_volt_low, sweep_volt_high))
    print("Power Supply Current: {}".format(psu_curr))
    print("Power ON/OFF Cycle @ Stepping: {}".format([' NO', 'YES'][pl_cycle_step]))
    print("Power Load Max Voltage: {}".format(pl_max_volt))
    print("Power Load Max Current: {}".format(pl_max_curr))
    print("Power Load Sweep Current: 0 to {} @ step {}".format(sweep_curr_max / 1000, sweep_curr_step / 1000))

    spd_hdr.spdSetup('CH1:VOLT {0:.2f}'.format(sweep_volt_low).encode())
    spd_hdr.spdSetup('CH1:CURR {0:.2f}'.format(psu_curr).encode())
    print('CH1:CURR {0:.2f}'.format(psu_curr).encode())

    print("Press Y to confirm!")
    ans = input()
    print("User Entered: {}".format(ans))
    if ans.lower()[0] is not 'y':
        pld_hdr.it85_close()
        spd_hdr.spdClose()
        exit()

    if not os.path.exists(os.path.join(os.path.curdir, filename)):
        print('WB not exist')
        wb = Workbook()
        wb.save(filename)
        wb.close()

    wb = load_workbook(filename)
    sh_name = wb.sheetnames
    if 'Sheet' in sh_name:
        ss = wb['Sheet']
        ss.title = 'ms_{}'.format(ts)
    else:
        wb.create_sheet('ms_{}'.format(ts))

    sheet = wb['ms_{}'.format(ts)]

    sheet.append([str(sys.version_info)])

    sheet.append([str(pld_hdr.it85_equipment_info())])
    pld_hdr.it85_get_settings()
    sheet.append([str(pld_hdr.it85_info_dict['System Version'])])

    sheet.append([str(spd_hdr.spdQuery(b'*IDN?'))])

    c1 = LineChart()
    c1.title = "Efficiency"
    c1.style = 13
    c1.y_axis.title = 'Power (W)'
    c1.x_axis.title = 'Load Current (A)'
    c1.height = 20
    c1.width = 30

    c2 = LineChart()
    c2.title = "Load Line Regulation"
    c2.style = 13
    c2.y_axis.title = 'Output Voltage (V)'
    c2.x_axis.title = 'Load Current (A)'
    c2.height = 20
    c2.width = 30

    c3 = LineChart()
    c3.title = "Load Line Regulation Error"
    c3.style = 13
    c3.y_axis.title = 'Normalized Error (%)'
    c3.x_axis.title = 'Load Current (A)'
    c3.height = 20
    c3.width = 30

    sweep_step_mv = int(sweep_step * 1000)
    sweep_volt_low_mv = int(sweep_volt_low * 1000)
    sweep_volt_high_mv = int(sweep_volt_high * 1000) + sweep_step_mv

    for j in range(sweep_volt_low_mv, sweep_volt_high_mv, sweep_step_mv):
        for k in range(0, repeat, 1):

            sheet.append([])

            psu_off(spd_hdr)
            pld_hdr.it85_channel_on_off(False) if pl_cycle_step else None
            pld_hdr.it85_cc_current(0.0)

            spd_hdr.spdSetup('CH1:VOLT {0:.2f}'.format(j / 1000).encode())
            spd_hdr.spdSetup('CH1:CURR {0:.2f}'.format(psu_curr).encode())

            psu_on(spd_hdr)
            # sleep(1.0)

            sheet.append([user_msg])

            sheet.append(['PSU Set Volt', 'PSU Set Curr',
                          'PS Volt', 'PS Curr', 'PS Pow',
                          'PL Volt', 'PL Curr', 'PL Pow'])

            data = [spd_hdr.spdQuery(b'CH1:VOLT?'), spd_hdr.spdQuery(b'CH1:CURR?')]

            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                future_pld = executor.submit(pld_acq, pld_hdr)
                future_psu = executor.submit(psu_acq, spd_hdr)
            prg_load = future_pld.result()
            psu_load = future_psu.result()
            data += psu_load
            data += prg_load

            sheet.append(data)

            # ===================================================================
            # Set PWR LD
            # ===================================================================
            pld_hdr.it85_cc_setup(pl_max_curr, pl_max_pwr, pl_4wire_sen, False, pl_max_volt)

            sheet.append([])
            sheet.append(['CC Set',
                          'PL Volt - V@{}'.format(j / 1000), 'PL Curr', 'PL Pow',
                          'PS Volt', 'PS Curr', 'PS Pow',
                          'Eff - V@{}'.format(j / 1000),
                          ''])

            start_idx_r = sheet.max_row
            start_idx_c = sheet.max_column
            cell_lut_volt = '{}{}'.format(
                get_column_letter(start_idx_c), start_idx_r)
            pwr_avg = []

            print('Start sweep CC loading @ VIN={}'.format(j / 1000))

            pld_hdr.it85_channel_on_off(True)

            for i in range(0, sweep_curr_max + sweep_curr_step, sweep_curr_step):

                data = [pld_hdr.it85_cc_current(i / 1000)]
                sleep(1.00)

                with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
                    future_pld = executor.submit(pld_acq, pld_hdr)
                    future_psu = executor.submit(psu_acq, spd_hdr)
                prg_load = future_pld.result()
                psu_load = future_psu.result()
                data += prg_load
                data += psu_load

                try:
                    pwr_cal = float(prg_load[-1]) / (float(psu_load[-1]) + offset) * 100.0
                    pwr_cal = pwr_cal if pwr_cal <= 100.0 else 100.0
                except ZeroDivisionError:
                    pwr_cal = 0.0
                pwr_avg.append(pwr_cal)
                print(
                    'Efficiency {:07.3f} @ VIN={:02.3f}, VOUT={:02f}, CC={:02.3f}, PIN={:07.3f}, POUT={:07.3f}'.format(
                        pwr_cal, j / 1000, float(prg_load[0]), i / 1000, float(psu_load[-1]) + offset,
                        float(prg_load[-1])))
                data += [pwr_cal]

                cell_pl_volt = '{}{}'.format(
                    get_column_letter(2), sheet.max_row + 1)
                load_line_formula = '=({}-{})/{}'.format(
                    cell_pl_volt, cell_lut_volt, cell_lut_volt)
                data += [load_line_formula]
                sheet.append(data)

            print('Efficiency Avg {:.3f}'.format(sum(pwr_avg) / len(pwr_avg)))
            print('Efficiency Max {:.3f}'.format(max(pwr_avg)))
            print('Efficiency Min {:.3f}'.format(min(pwr_avg)))

            end_idx_r = sheet.max_row

            # insert load and line regulation
            sheet[cell_lut_volt] = '=VLOOKUP({}{},{}{}:{}{},2)'.format(
                get_column_letter(1), start_idx_r + int((end_idx_r - start_idx_r) / 2) + 1,
                get_column_letter(1), start_idx_r + 1,
                get_column_letter(2), end_idx_r)

            data = Reference(sheet,
                             min_col=start_idx_c - 1, min_row=start_idx_r,
                             max_col=start_idx_c - 1, max_row=end_idx_r)
            label = Reference(sheet,
                              min_col=1, min_row=start_idx_r + 1,
                              max_col=1, max_row=end_idx_r)
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(label)
            c1.y_axis.scaling.min = 0
            c1.y_axis.scaling.max = 100

            data = Reference(sheet,
                             min_col=start_idx_c - 7, min_row=start_idx_r,
                             max_col=start_idx_c - 7, max_row=end_idx_r)
            label = Reference(sheet,
                              min_col=1, min_row=start_idx_r + 1,
                              max_col=1, max_row=end_idx_r)
            c2.add_data(data, titles_from_data=True)
            c2.set_categories(label)

            data = Reference(sheet,
                             min_col=start_idx_c, min_row=start_idx_r + 1,
                             max_col=start_idx_c, max_row=end_idx_r)
            label = Reference(sheet,
                              min_col=1, min_row=start_idx_r + 1,
                              max_col=1, max_row=end_idx_r)
            s = Series(values=data, title='Line V={}'.format(j / 1000))
            c3.series.append(s)
            c3.set_categories(label)

            psu_off(spd_hdr)
            pld_hdr.it85_channel_on_off(False)

    sheet.add_chart(c1, get_column_letter(sheet.max_column + 3) + str(1))
    sheet.add_chart(c2, get_column_letter(sheet.max_column + 3) + str(c1.height * 2 + 1))
    sheet.add_chart(c3, get_column_letter(sheet.max_column + 3) + str(c1.height * 4 + 1))

    pld_hdr.it85_close()
    spd_hdr.spdClose()

    while True:
        try:
            wb.save(filename=filename)
            break
        except PermissionError:
            print('Please close file and enter y[Y] and cont\'d')
            ans = input()
            print("User Entered: {}".format(ans))
            if ans.lower()[0] is not 'y':
                exit()
    print("Measurement is DONE, Raw Data Saved to {}.".format(filename))


if __name__ == '__main__':
    main()
