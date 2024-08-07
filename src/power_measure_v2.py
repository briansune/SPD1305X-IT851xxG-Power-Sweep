import spd1000_series as _spd
import math
from datetime import datetime
import serial
import serial.tools.list_ports
from time import sleep
import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import os

sys_ver = 0
serial_holder = serial.Serial()


def port2list():
    l_ports = serial.tools.list_ports.comports()
    result = [element.device for element in l_ports]
    return result


def sh_wr(tx_str):
    # print(sys_ver)
    serial_holder.write(tx_str.encode() if (sys_ver >= 3) else tx_str)


def sh_rd():
    # print(sys_ver)
    s = serial_holder.read_until()
    s = s.decode().strip() if (sys_ver >= 3) else s.strip()
    if s[0] is 'R':
        s = s[1:].lstrip()

    return s


def psu_on(spd_hdr):
    while True:
        spd_hdr.spdSetup(b'OUTP CH1,ON')
        res = spd_hdr.spdQuery(b'SYST:STAT?')
        if int(res, 16) & 0x10 == 0x10:
            break


def psu_off(spd_hdr):
    while True:
        spd_hdr.spdSetup(b'OUTP CH1,OFF')
        res = spd_hdr.spdQuery(b'SYST:STAT?')
        if int(res, 16) & 0x10 == 0x00:
            break


def main():
    spd_hdr = _spd.SPD1000()
    spd_hdr.spd_ip = "192.168.1.214"

    global sys_ver
    global serial_holder

    sys_ver = sys.version_info[0]
    print(sys_ver)
    print(port2list())

    serial_holder = serial.Serial('COM23', 115200, timeout=1)
    # serial_holder = serial.Serial('COM12', 115200, timeout=1)
    # serial_holder = serial.Serial('COM4', 115200, timeout=1)
    port_status = serial_holder.isOpen()
    if port_status:
        serial_holder.close()

    serial_holder.open()
    spd_hdr.spdConnect()

    dt = datetime.now()
    ts = math.floor(datetime.timestamp(dt))

    sh_wr('CH:SW OFF\x0A')
    sh_rd()
    sh_wr('CH:SW?\x0A')
    sh_rd()

    # ==============================================================================
    # Change the following parameters!
    # ==============================================================================
    spd_hdr.spdSetup(b'*UNLOCK')
    # spd_hdr.spdSetup(b'MODE:SET 2W')
    spd_hdr.spdSetup(b'MODE:SET 4W')
    sweep_volt_low = 26
    sweep_volt_high = 30
    # Repeat Num set
    repeat = 1
    # Other offset power
    offset = 0.0
    pl_max_curr = 21.5
    pl_max_volt = 4.5
    sweep_curr_max = 2000
    sweep_curr_step = 20
    psu_curr = (sweep_curr_max / 100 * pl_max_volt) / sweep_volt_low * 1.2
    psu_curr = 5.0 if psu_curr > 5.0 else psu_curr
    
    filename = f'TI_LM25116_27to30V_WEBENCH_T3.xlsx'
    print(filename)

    print("Power Supply Voltage Range: {} to {}".format(sweep_volt_low, sweep_volt_high))
    print("Power Supply Current: {}".format(psu_curr))
    print("Power Load Max Voltage: {}".format(pl_max_volt))
    print("Power Load Max Current: {}".format(pl_max_curr))
    print("Power Load Sweep Current: 0 to {} @ step {}".format(sweep_curr_max / 100, sweep_curr_step / 100))

    spd_hdr.spdSetup('CH1:VOLT {0:.2f}'.format(sweep_volt_low).encode())
    spd_hdr.spdSetup('CH1:CURR {0:.2f}'.format(psu_curr).encode())
    print('CH1:CURR {0:.2f}'.format(psu_curr).encode())

    ans = input("Press Y to confirm!\n")
    if ans.lower()[0] is not 'y':
        serial_holder.close()
        spd_hdr.spdClose()
        exit()

    if not os.path.exists(os.path.join(os.path.curdir, filename)):
        print('WB not exist')
        wb = Workbook()
        wb.save(filename)
        wb.close()

    wb = load_workbook(filename)
    sh_name = wb.sheetnames
    if 'Sheet' in sh_name:
        ss = wb['Sheet']
        ss.title = 'ms_{}'.format(ts)
    else:
        wb.create_sheet('ms_{}'.format(ts))

    sheet = wb['ms_{}'.format(ts)]

    sheet.append([str(sys.version_info)])
    sh_wr('*IDN?\x0A')
    sheet.append([str(sh_rd())])
    sh_wr('SYSTem:VERSion?\x0A')
    sheet.append([str(sh_rd())])
    sheet.append([str(spd_hdr.spdQuery(b'*IDN?'))])

    c1 = LineChart()
    c1.title = "Efficiency"
    c1.style = 13
    c1.y_axis.title = 'Power (W)'
    c1.x_axis.title = 'Load Current (A)'
    c1.height = 20
    c1.width = 30

    for j in range(sweep_volt_low, sweep_volt_high + 1, 1):
        for k in range(0, repeat, 1):

            sheet.append([])

            psu_off(spd_hdr)
            sh_wr('CH:SW OFF\x0A')
            sh_rd()
            sh_wr('CURR:CC {0:.3f}\x0A'.format(0))
            sh_rd()

            spd_hdr.spdSetup('CH1:VOLT {0:.2f}'.format(j).encode())
            spd_hdr.spdSetup('CH1:CURR {0:.2f}'.format(psu_curr).encode())

            psu_on(spd_hdr)
            sleep(3.0)

            sheet.append(['PSU Set Volt', 'PSU Set Curr',
                          'PS Volt', 'PS Curr', 'PS Pow',
                          'PL Volt', 'PL Curr', 'PL Pow'])

            data = [spd_hdr.spdQuery(b'CH1:VOLT?'), spd_hdr.spdQuery(b'CH1:CURR?')]

            data += [spd_hdr.spdQuery(b'MEAS:VOLT?'), spd_hdr.spdQuery(b'MEAS:CURR?'),
                     spd_hdr.spdQuery(b'MEAS:POWE?')]

            prg_load = []
            sh_wr('MEAS:VOLT?\x0A')
            prg_load.append(sh_rd())
            sh_wr('MEAS:CURR?\x0A')
            prg_load.append(sh_rd())
            sh_wr('MEAS:POWE?\x0A')
            prg_load.append(sh_rd())
            data += prg_load
            sheet.append(data)

            # ===================================================================
            # Set PWR LD
            # ===================================================================
            sh_wr('CURR:IMAX {:.2f}\x0A'.format(pl_max_curr))
            sh_rd()
            sh_wr('VOLT:VMAX {:.2f}\x0A'.format(pl_max_volt))
            sh_rd()

            sheet.append([])
            sheet.append(['CC Set',
                          'PL Volt', 'PL Curr', 'PL Pow',
                          'PS Volt', 'PS Curr', 'PS Pow',
                          'Eff - V@{}'.format(j)])

            start_idx_r = sheet.max_row
            start_idx_c = sheet.max_column
            pwr_avg = []

            print('Start sweep CC loading @ VIN={}'.format(j))

            sh_wr('CH:SW ON\x0A')
            sh_rd()

            for i in range(0, sweep_curr_max + sweep_curr_step, sweep_curr_step):

                sh_wr('CURR:CC {0:.3f}\x0A'.format(i / 100))
                sh_rd()
                sh_wr('CURR:CC?\x0A')
                data = [sh_rd()]

                sleep(0.4)

                prg_load = []
                sh_wr('MEAS:VOLT?\x0A')
                prg_load.append(sh_rd())
                sh_wr('MEAS:CURR?\x0A')
                prg_load.append(sh_rd())
                sh_wr('MEAS:POWE?\x0A')
                prg_load.append(sh_rd())
                data += prg_load

                psu_load = [spd_hdr.spdQuery(b'MEAS:VOLT?'), spd_hdr.spdQuery(b'MEAS:CURR?'),
                            spd_hdr.spdQuery(b'MEAS:POWE?')]
                data += psu_load
                try:
                    pwr_cal = float(prg_load[-1]) / (float(psu_load[-1]) + offset) * 100.0
                    pwr_cal = pwr_cal if pwr_cal <= 100.0 else 100.0
                except ZeroDivisionError:
                    pwr_cal = 0.0
                pwr_avg.append(pwr_cal)
                print('Efficiency {:07.3f} @ VIN={:02d}, VOUT={:02f}, CC={:03f}, PIN={:07.3f}, POUT={:07.3f}'.format(
                    pwr_cal, j, float(prg_load[0]), i / 100, float(psu_load[-1]) + offset, float(prg_load[-1])))
                data += [pwr_cal]
                sheet.append(data)

            print('Efficiency Avg {:.3f}'.format(sum(pwr_avg) / len(pwr_avg)))

            end_idx_r = sheet.max_row
            data = Reference(sheet,
                             min_col=start_idx_c, min_row=start_idx_r,
                             max_col=start_idx_c, max_row=end_idx_r)
            label = Reference(sheet,
                              min_col=1, min_row=start_idx_r + 1,
                              max_col=1, max_row=end_idx_r)
            c1.add_data(data, titles_from_data=True)
            c1.set_categories(label)

            psu_off(spd_hdr)
            sh_wr('CH:SW OFF\x0A')
            sh_rd()
            sleep(0.05)

    sheet.add_chart(c1, get_column_letter(sheet.max_column + 3) + str(1))

    serial_holder.close()
    spd_hdr.spdClose()

    wb.save(filename=filename)


if __name__ == '__main__':
    main()
